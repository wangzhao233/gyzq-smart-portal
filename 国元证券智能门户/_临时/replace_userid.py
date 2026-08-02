"""国元证券通讯录 - F_USERID 替换为 USERID"""
from openpyxl import load_workbook
from pathlib import Path

SRC = Path(r"C:\Users\11039\Desktop\国元证券通讯录.xlsx")
MAP = Path(r"C:\Users\11039\Desktop\a.xlsx")
OUT = SRC.parent / "国元证券通讯录_已替换USERID.xlsx"

# 1. 加载映射表: F_USERID → USERID
print("加载映射表...")
wb_map = load_workbook(MAP, read_only=True)
ws_map = wb_map.active
mapping = {}
for i, row in enumerate(ws_map.iter_rows(min_row=2, values_only=True)):  # 跳过表头
    f_userid, userid, _ = row
    if f_userid and userid:
        mapping[f_userid] = str(userid)
wb_map.close()
print(f"  映射条目: {len(mapping)}")

# 2. 加载通讯录
print("加载通讯录...")
wb = load_workbook(SRC)
ws = wb.active
print(f"  Sheet: {ws.title}, 行数: {ws.max_row}")

# 3. 替换 B 列 (索引=2, 即第2列)
replaced = 0
missing = 0
for row in range(2, ws.max_row + 1):  # 从第2行开始（跳过表头）
    cell = ws.cell(row=row, column=2)
    f_userid = str(cell.value).strip() if cell.value else ""
    if f_userid in mapping:
        cell.value = mapping[f_userid]
        replaced += 1
    else:
        missing += 1

# 更新表头
ws.cell(row=1, column=2).value = "USERID"

# 删除 C 列残留的 VLOOKUP 公式（如果存在）
c_header = ws.cell(row=1, column=3).value
c_first = ws.cell(row=2, column=3).value
if c_header == "USERID" and c_first and "VLOOKUP" in str(c_first):
    ws.delete_cols(3)  # 删除第3列
    print("  已删除 C 列残留 VLOOKUP 公式")
elif c_header == "USERID":
    ws.delete_cols(3)
    print("  已删除 C 列（与B列重复的USERID列）")

# 4. 保存
print(f"保存结果: {OUT}")
wb.save(OUT)
wb.close()

print(f"\n完成！替换 {replaced} 条, 未找到映射 {missing} 条")
if missing:
    print("未找到映射的可能是系统账号(如admin)或已删除用户，保留原值")

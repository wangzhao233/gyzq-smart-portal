# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "第三方应用接入跟进表"

# 表头样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="4472C4")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# 分类标题样式
category_font = Font(bold=True, size=12, color="FFFFFF")
category_fill = PatternFill("solid", fgColor="2F5496")

# 数据样式
data_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(size=10)

# 边框
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 状态颜色
status_fills = {
    "已部署": PatternFill("solid", fgColor="C6EFCE"),
    "已验证": PatternFill("solid", fgColor="C6EFCE"),
    "待迁移": PatternFill("solid", fgColor="FFEB9C"),
    "待验证": PatternFill("solid", fgColor="F2F2F2"),
}

# 设置列宽
col_widths = {'A': 6, 'B': 20, 'C': 15, 'D': 25, 'E': 20, 'F': 15, 'G': 15, 'H': 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# 标题行
ws.merge_cells('A1:H1')
title_cell = ws['A1']
title_cell.value = "第三方应用接入跟进表 - 国元证券企微私有化应用迁移"
title_cell.font = Font(bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="1F4E79")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# 副标题
ws.merge_cells('A2:H2')
sub_cell = ws['A2']
sub_cell.value = "更新日期：2026-06-24 | 数据来源：葛老师提供 | 项目：国元证券企微私有化"
sub_cell.font = Font(size=10, italic=True)
sub_cell.alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 20

# 定义数据
headers = ["序号", "应用名称", "应用编码", "归属团队", "联系人", "迁移日期/对接人", "状态", "备注"]

# 新建应用数据
new_apps = [
    [1, "智能门户", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "已部署", "信息汇聚"],
    [2, "统一待办中心", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "已部署", "待办/消息汇聚"],
    [3, "邮件", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "已部署", "原生体验"],
    [4, "日程中心", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "已部署", "日程汇聚"],
    [5, "私有版原生应用", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "已部署", "IM、文档、微盘、收集表"],
    [6, "企业小助手", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "待迁移", "在私有版企微上重建"],
    [7, "考勤打卡", "-", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "待迁移", "在私有版企微上重建"],
]

# 迁移-无需修改数据
migrate_no_change = [
    [8, "兴趣社团", "1000013", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [9, "论党的青年工作", "1000015", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [10, "投教基地", "1000017", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [11, "国元荣誉", "1000018", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [12, "每周食谱", "1000024", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "待验证", ""],
    [13, "出差申请", "1000044", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "待验证", ""],
    [14, "国元邮箱", "1000065", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [15, "旧版投行", "1000067", "信息技术部/管理系统组", "陈洋、何洋", "-", "待验证", ""],
    [16, "数聚驾驶舱", "1000068", "金融科技部/数据开发组", "闫天翔、李铎", "6/17 已验证", "已验证", ""],
    [17, "快递收发", "1000097", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [18, "洗衣会员卡", "1000099", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [19, "国元风采", "1000114", "金融科技部/应用开发组", "陈洋、何洋、李晓鹤", "-", "待验证", ""],
    [20, "诵歌二十大", "1000115", "金融科技部/应用开发组", "陈洋、何洋", "-", "待验证", ""],
    [21, "重点工作督办", "1000135", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "待验证", ""],
    [22, "合规智库", "1000161", "金融科技部/管理开发组", "程文斐、费晓璐", "-", "待验证", ""],
    [23, "商务招待申请", "1000169", "金融科技部/管理开发组", "程文斐", "-", "待验证", ""],
    [24, "网球场地预约", "1000175", "金融科技部/管理开发组", "程文斐、葛鹏飞", "-", "待验证", ""],
]

# 迁移-修改配置参数数据
migrate_config = [
    [25, "衡泰移动审批", "1000012", "信息技术部/业务支持组", "陈辉", "待定", "待迁移", ""],
    [26, "数据摆渡", "1000033", "信息技术部/基础平台组", "李振", "待定", "待迁移", ""],
    [27, "商旅平台", "1000045", "信息技术部/管理系统组", "王文枫", "6/10 陈俊卿", "已验证", ""],
    [28, "国元E学", "1000066", "信息技术部/管理系统组", "方明、阙宏伟", "6/22 方明", "待迁移", ""],
    [29, "意见建议", "1000087", "金融科技部/管理开发组", "程文斐、葛鹏飞", "待定", "待迁移", ""],
    [30, "新版ERP", "1000112", "金融科技部/管理开发组", "程文斐、葛鹏飞", "待定", "待迁移", ""],
    [31, "投行综合", "1000123", "信息技术部/管理系统组", "戈明东", "已排期 王迎池", "待迁移", ""],
    [32, "智慧投行", "1000125", "信息技术部/管理系统组", "戈明东、张亚飞、陈龙", "已排期 王迎池", "待迁移", ""],
    [33, "新统一认证", "1000129", "金融科技部/管理开发组", "程文斐、葛鹏飞", "待定", "待迁移", ""],
    [34, "薪酬管理", "1000131", "金融科技部/管理开发组", "程文斐、葛鹏飞", "待定", "待迁移", ""],
    [35, "管理会计", "1000134", "信息技术部/管理系统组", "王文枫", "已排期 张宇豪", "待迁移", ""],
    [36, "燎元智能助手", "1000137", "金融科技部/人工智能组", "高帆、何洋", "待定", "待迁移", ""],
    [37, "融和", "1000139", "信息技术部/移动互联组", "孙靖明", "已排期 王美玉", "待迁移", "开发商提出改造费用"],
    [38, "资讯中心监控", "1000145", "金融科技部/投资科技组", "王雪峰", "6/18 王雪峰", "待迁移", ""],
    [39, "财务共享报销", "1000146", "信息技术部/管理系统组", "王文枫", "待定", "待迁移", ""],
    [40, "审计管理", "1000149", "信息技术部/管理系统组", "戈明东、陈俊卿", "已排期 陈俊卿", "待迁移", ""],
    [41, "绩效管理", "1000152", "金融科技部/管理开发组", "程文斐、葛鹏飞", "已排期 张宇豪", "待迁移", ""],
    [42, "战略与目标管理", "1000154", "金融科技部/管理开发组", "程文斐、葛鹏飞", "待定", "待迁移", ""],
    [43, "元心助手", "1000158", "金融科技部/人工智能组", "高帆、何洋", "待定", "待迁移", ""],
    [44, "合同管理", "1000162", "信息技术部/管理系统组", "戈明东、费晓璐", "已排期 王迎池", "待迁移", ""],
    [45, "元心基层风采展示", "1000163", "金融科技部/管理开发组", "程文斐", "待定", "待迁移", ""],
    [46, "融汇平台", "1000165", "金融科技部/投资科技组", "杨洪军", "6/15 邸云龙", "已验证", ""],
    [47, "AI能力平台", "1000173", "金融科技部/人工智能组", "杨鹏、何洋、李晓鹤", "待定", "待迁移", ""],
]

# 迁移-重新扫码绑定数据
migrate_rescan = [
    [48, "尽调采集助手", "1000157", "信息技术部/管理系统组", "戈明东、周大农", "-", "待迁移", "小程序重新扫码"],
    [49, "电子名片", "1000160", "金融科技部/管理开发组", "蒋宇澄、张磊磊、杨容季", "-", "待迁移", "小程序重新扫码"],
]

# 统一待办中心对接计划
todo_plan = [
    ["智慧投行", "1000125", "王迎池", "已排期", ""],
    ["管理会计", "1000134", "张宇豪", "已排期", ""],
    ["融和", "1000139", "王美玉", "待定", "开发商提出改造费用"],
    ["审计管理", "1000149", "陈俊卿", "已排期", ""],
    ["绩效管理", "1000152", "张宇豪", "已排期", ""],
]

# 写入分类标题和数据
current_row = 4

# 一、新建应用（7个）
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="一、新建应用（7个 - 已部署）")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

# 写入表头
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

# 写入新建应用数据
for row_data in new_apps:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if col == 7 and value in status_fills:
            cell.fill = status_fills[value]
    ws.row_dimensions[current_row].height = 18
    current_row += 1

current_row += 1

# 二、迁移应用 - 无需修改（17个）
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="二、迁移应用 - 无需修改（17个）")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

for row_data in migrate_no_change:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if col == 7 and value in status_fills:
            cell.fill = status_fills[value]
    ws.row_dimensions[current_row].height = 18
    current_row += 1

current_row += 1

# 三、迁移应用 - 修改配置参数（23个）
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="三、迁移应用 - 修改配置参数（23个）")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

for row_data in migrate_config:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if col == 7 and value in status_fills:
            cell.fill = status_fills[value]
    ws.row_dimensions[current_row].height = 18
    current_row += 1

current_row += 1

# 四、迁移应用 - 重新扫码绑定（2个）
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="四、迁移应用 - 重新扫码绑定（2个）")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

for row_data in migrate_rescan:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if col == 7 and value in status_fills:
            cell.fill = status_fills[value]
    ws.row_dimensions[current_row].height = 18
    current_row += 1

current_row += 1

# 五、统一待办中心对接计划
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="五、统一待办中心对接计划（5个应用）")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

todo_headers = ["应用名称", "应用编码", "对接人", "预计日期", "备注"]
for col, header in enumerate(todo_headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

for row_data in todo_plan:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
    ws.row_dimensions[current_row].height = 18
    current_row += 1

current_row += 2

# 六、进度看板
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="六、进度看板（更新至2026-06-24）")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

dashboard_headers = ["状态", "数量", "说明"]
for col, header in enumerate(dashboard_headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

dashboard_data = [
    ["已验证", 3, "商旅平台(6/10)、融汇平台(6/15)、数聚驾驶舱(6/17)"],
    ["待迁移（已排期）", 8, "资讯中心监控(6/18)、国元E学(6/22)等"],
    ["待排期", 29, "联系方式待补充"],
    ["小程序扫码重绑", 2, "尽调采集助手、电子名片"],
    ["合计", 42, ""],
]

for row_data in dashboard_data:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if col == 1 and value in status_fills:
            cell.fill = status_fills[value]
    ws.row_dimensions[current_row].height = 18
    current_row += 1

current_row += 1

# 七、风险与阻塞项
ws.merge_cells(f'A{current_row}:H{current_row}')
cat_cell = ws.cell(row=current_row, column=1, value="七、风险与阻塞项")
cat_cell.font = category_font
cat_cell.fill = category_fill
cat_cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[current_row].height = 25
current_row += 1

risk_headers = ["风险", "涉及应用", "说明", "优先级"]
for col, header in enumerate(risk_headers, 1):
    cell = ws.cell(row=current_row, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[current_row].height = 20
current_row += 1

risk_data = [
    ["开发商改造费用", "融和(1000139)", "建议在企微侧做新老接口适配中转", "高"],
    ["联系人缺失", "多个", "部分应用迁移日期和对接人待补充", "中"],
    ["应用对接延期风险", "29个待排期应用", "计划6/25前完成压力较大", "高"],
    ["组织架构同步", "人员组织机构", "登录已通过，组织架构同步开发中", "中"],
    ["待办中心对接", "5个应用", "第三方应用对接待办中心联调测试有风险", "中"],
]

risk_fills = {
    "高": PatternFill("solid", fgColor="FFC7CE"),
    "中": PatternFill("solid", fgColor="FFEB9C"),
}

for row_data in risk_data:
    for col, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col, value=value)
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        if col == 4 and value in risk_fills:
            cell.fill = risk_fills[value]
    ws.row_dimensions[current_row].height = 18
    current_row += 1

# 冻结窗格
ws.freeze_panes = 'A3'

# 保存
output_path = r"C:\Users\11039\.workbuddy\workspace\files\30712\0c43f5b1-6269-4760-8a5f-bb5b29a666cc\第三方应用接入跟进表_整合版.xlsx"
wb.save(output_path)
print(f"Excel已保存: {output_path}")
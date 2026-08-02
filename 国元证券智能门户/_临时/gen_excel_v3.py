from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "应用对接进度"

# ── Colors ──
DARK_BLUE = "1F3864"
MED_BLUE = "2F5496"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
GREEN_FILL = "E2EFDA"
GREEN_FONT = "375623"
YELLOW_FILL = "FFF2CC"
YELLOW_FONT = "806000"
RED_FILL = "FCE4D6"
HEADER_BG = DARK_BLUE
SUBHEADER_BG = MED_BLUE

# ── Styles ──
thin = Side(style='thin', color='B4C6E7')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

title_font = Font(name='微软雅黑', bold=True, size=14, color=DARK_BLUE)
subtitle_font = Font(name='微软雅黑', size=10, color='808080')
header_font = Font(name='微软雅黑', bold=True, size=10, color=WHITE)
body_font = Font(name='微软雅黑', size=10, color='333333')
done_font = Font(name='微软雅黑', size=10, color=GREEN_FONT, bold=True)
pending_font = Font(name='微软雅黑', size=10, color=YELLOW_FONT)
section_font = Font(name='微软雅黑', bold=True, size=11, color=WHITE)

# ── Title area ──
ws.merge_cells('A1:H1')
c = ws['A1']
c.value = '国元证券 · 企业微信私有化 · 第三方应用对接进度表'
c.font = title_font
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 36

ws.merge_cells('A2:H2')
c = ws['A2']
c.value = '更新日期：2026年6月24日  |  数据来源：客户系统实际截图确认  |  合计 67 个应用'
c.font = subtitle_font
c.alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[2].height = 22

row = 4

# ── Summary cards at top ──
cards = [
    ('已完成', '55 个', '已在客户系统中确认存在', GREEN_FILL, GREEN_FONT),
    ('待完成', '12 个', '截图未覆盖，需进一步确认', YELLOW_FILL, YELLOW_FONT),
    ('合计', '67 个', '含 7 个新建 + 42 个迁移 + 18 个新发现', LIGHT_BLUE, DARK_BLUE),
]
for i, (label, num, desc, bg, fc) in enumerate(cards):
    col_start = i * 3 + 1
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row+1, end_column=col_start+1)
    c = ws.cell(row=row, column=col_start, value=label)
    c.font = Font(name='微软雅黑', bold=True, size=12, color=fc)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border
    for rr in [row, row+1]:
        for cc in [col_start, col_start+1]:
            cell = ws.cell(row=rr, column=cc)
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            cell.border = border

    ws.merge_cells(start_row=row, start_column=col_start+2, end_row=row, end_column=col_start+2)
    c = ws.cell(row=row, column=col_start+2, value=num)
    c.font = Font(name='微软雅黑', bold=True, size=18, color=fc)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

    c = ws.cell(row=row+1, column=col_start+2, value=desc)
    c.font = Font(name='微软雅黑', size=9, color=fc)
    c.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = border

row += 3

# ── Column headers ──
headers = ['序号', '应用名称', '应用编码', '归属部门', '联系人', '计划/完成日期', '状态', '备注']
col_widths = [6, 24, 12, 30, 24, 18, 10, 28]
for col, (h, w) in enumerate(zip(headers, col_widths), 1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = header_font
    c.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type='solid')
    c.alignment = center
    c.border = border
    ws.column_dimensions[get_column_letter(col)].width = w
ws.row_dimensions[row].height = 28
row += 1

# ── Data ──
apps = [
    # (序号, 名称, 编码, 部门, 联系人, 日期, 状态, 备注)
    # ── 新建应用 ──
    ('SECTION', '一、新建应用（7个）'),
    (1, '智能门户', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', '信息汇聚门户'),
    (2, '统一待办', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', '待办消息汇聚'),
    (3, '日程中心', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', '日程汇聚'),
    (4, '邮箱提醒', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', '原生邮件体验'),
    (5, '企业小助手', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', '需在私有版企微重建'),
    (6, '考勤打卡', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', '需在私有版企微重建'),
    (7, '私有版原生应用', '-', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', 'IM / 文档 / 微盘 / 收集表'),

    # ── 迁移应用 ──
    ('SECTION', '二、迁移应用（42个）— 其中30个已完成、12个待完成'),
    (8, '兴趣社团', '1000013', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (9, '论党的青年工作', '1000015', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (10, '投教基地', '1000017', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (11, '国元荣誉', '1000018', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (12, '国元风采', '1000114', '金融科技部/应用开发组', '陈洋、何洋、李晓鹤', '-', '已完成', ''),
    (13, '颂歌二十大', '1000115', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (14, '国元邮箱', '1000065', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (15, '快递收发', '1000097', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (16, '洗衣会员卡', '1000099', '金融科技部/应用开发组', '陈洋、何洋', '-', '已完成', ''),
    (17, '合规智库', '1000161', '金融科技部/管理开发组', '程文斐、费晓璐', '-', '已完成', ''),
    (18, '出差申请', '1000044', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', ''),
    (19, '数聚驾驶舱', '1000068', '金融科技部/数据开发组', '闫天翔、李铎', '2026-06-17', '已完成', '6/17 已验证'),
    (20, '衡泰移动审批', '1000012', '信息技术部/业务支持组', '陈辉', '-', '已完成', ''),
    (21, '商旅平台', '1000045', '信息技术部/管理系统组', '王文枫', '2026-06-10', '已完成', '6/10 陈俊卿 已验证'),
    (22, '国元E学', '1000066', '信息技术部/管理系统组', '方明、阙宏伟', '2026-06-22', '已完成', '对接人：方明'),
    (23, '意见建议', '1000087', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', ''),
    (24, '智慧投行', '1000125', '信息技术部/管理系统组', '戈明东、张亚飞、陈龙', '已排期', '已完成', '对接人：王迎池'),
    (25, '统一认证', '1000129', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '已完成', ''),
    (26, '管理会计', '1000134', '信息技术部/管理系统组', '王文枫', '已排期', '已完成', '对接人：张宇豪'),
    (27, '燎元智能助手', '1000137', '金融科技部/人工智能组', '高帆、何洋', '-', '已完成', ''),
    (28, '融和平台', '1000139', '信息技术部/移动互联组', '孙靖明', '已排期', '已完成', '对接人：王美玉；开发商提出改造费用'),
    (29, '资讯中心监控', '1000145', '金融科技部/投资科技组', '王雪峰', '2026-06-18', '已完成', '6/18 王雪峰'),
    (30, '财务共享报销', '1000146', '信息技术部/管理系统组', '王文枫', '-', '已完成', '截图显示「财务共享」'),
    (31, '审计管理', '1000149', '信息技术部/管理系统组', '戈明东、陈俊卿', '已排期', '已完成', '对接人：陈俊卿'),
    (32, '合同管理', '1000162', '信息技术部/管理系统组', '戈明东、费晓璐', '已排期', '已完成', '对接人：王迎池'),
    (33, '元心基层风采展示', '1000163', '金融科技部/管理开发组', '程文斐', '-', '已完成', ''),
    (34, '融汇平台', '1000165', '金融科技部/投资科技组', '杨洪军', '2026-06-15', '已完成', '6/15 邸云龙 已验证'),

    # 待完成
    (35, '旧版投行', '1000067', '信息技术部/管理系统组', '陈洋、何洋', '-', '待完成', ''),
    (36, '重点工作督办', '1000135', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '待完成', ''),
    (37, '商务招待申请', '1000169', '金融科技部/管理开发组', '程文斐', '-', '待完成', ''),
    (38, '网球场地预约', '1000175', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '待完成', ''),
    (39, '每周食谱', '1000024', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '待完成', ''),
    (40, '数据摆渡', '1000033', '信息技术部/基础平台组', '李振', '-', '待完成', ''),
    (41, '新版ERP', '1000112', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '待完成', ''),
    (42, '投行综合', '1000123', '信息技术部/管理系统组', '戈明东', '已排期', '待完成', '对接人：王迎池'),
    (43, '薪酬管理', '1000131', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '待完成', ''),
    (44, '绩效管理', '1000152', '金融科技部/管理开发组', '程文斐、葛鹏飞', '已排期', '待完成', '对接人：张宇豪'),
    (45, '战略与目标管理', '1000154', '金融科技部/管理开发组', '程文斐、葛鹏飞', '-', '待完成', ''),
    (46, '元心助手', '1000158', '金融科技部/人工智能组', '高帆、何洋', '-', '待完成', ''),
    (47, 'AI能力平台', '1000173', '金融科技部/人工智能组', '杨鹏、何洋、李晓鹤', '-', '待完成', ''),
    (48, '国元尽调采集助手', '1000157', '信息技术部/管理系统组', '戈明东、周大农', '-', '已完成', '小程序'),
    (49, '国元证券电子名片', '1000160', '金融科技部/管理开发组', '蒋宇澄、张磊磊、杨容季', '-', '已完成', '小程序'),

    # ── 新发现 ──
    ('SECTION', '三、客户系统中新发现应用（18个，已完成）'),
    (50, '微盘扩容申请', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (51, '新闻公告', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (52, '快捷表单', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (53, '会议管家', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (54, '七巧plus', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (55, '七巧plus管理端', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (56, '智能门户管理端', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (57, '云办公_通讯录', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (58, '云办公_表单流程', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (59, '云办公_新闻公告', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (60, '移动外勤', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (61, '请假出差', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (62, 'SSO-TEST', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (63, '企微运维平台', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (64, '合规效能平台', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (65, '公告审批', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (66, '元心党建', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
    (67, '投行待办', '-', '待确认', '-', '-', '已完成', '客户系统发现'),
]

# ── Write rows ──
for item in apps:
    if item[0] == 'SECTION':
        # Section header row
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=item[1])
        c.font = section_font
        c.fill = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type='solid')
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = border
        for cc in range(2, 9):
            cell = ws.cell(row=row, column=cc)
            cell.fill = PatternFill(start_color=SUBHEADER_BG, end_color=SUBHEADER_BG, fill_type='solid')
            cell.border = border
        ws.row_dimensions[row].height = 26
        row += 1
        continue

    seq, name, code, dept, contact, date, status, note = item
    values = [seq, name, code, dept, contact, date, status, note]

    is_done = (status == '已完成')

    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.border = border
        c.alignment = center if col in [1, 6, 7] else left_align

        if is_done:
            c.fill = PatternFill(start_color=GREEN_FILL, end_color=GREEN_FILL, fill_type='solid')
            c.font = done_font
        elif status == '待完成':
            c.fill = PatternFill(start_color=YELLOW_FILL, end_color=YELLOW_FILL, fill_type='solid')
            c.font = pending_font
        else:
            c.font = body_font

        # Bold for app name column
        if col == 2:
            if is_done:
                pass  # already bold via done_font
            else:
                c.font = Font(name='微软雅黑', size=10, color='333333', bold=True)

    ws.row_dimensions[row].height = 22
    row += 1

# ── Freeze panes ──
ws.freeze_panes = 'A8'

# ── Print settings ──
ws.sheet_properties.pageSetUpPr = None
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_margins.left = 0.4
ws.page_margins.right = 0.4

filepath = '国元证券-应用对接进度表.xlsx'
wb.save(filepath)
print(f'OK: {filepath}')
